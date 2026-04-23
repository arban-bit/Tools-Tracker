import openpyxl
import os

path = r'c:\Users\arban\OneDrive - Vestas Wind Systems A S\_Offshore Readiness\Tools Tracker\V236 Nacelle Toolkit.xlsx'

if not os.path.exists(path):
    print(f'File not found: {path}')
else:
    wb = openpyxl.load_workbook(path, data_only=True)
    print(f'Sheets: {wb.sheetnames}')
    total_items = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f'\n=== {sheet_name} ===')
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            print(f'Headers: {rows[0]}')
            data_rows = rows[1:]
            for row in data_rows:
                if any(cell is not None for cell in row):
                    print(row)
                    total_items += 1
    print(f'\nTotal number of tools/items: {total_items}')
